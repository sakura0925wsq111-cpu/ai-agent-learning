"""Adapter contract plus separated HTTP, file, validation, and persistence helpers."""

from __future__ import annotations

import csv
import hashlib
import mimetypes
import re
import shutil
from abc import ABC, abstractmethod
from dataclasses import dataclass
from datetime import UTC, datetime
from html import unescape
from pathlib import Path
from time import sleep
from typing import Any
from urllib.parse import urlparse

import httpx
import pdfplumber
import xlrd
from openpyxl import load_workbook

from career_data.config import DATA_ROOT, HTTP_MAX_RETRIES, HTTP_TIMEOUT_SECONDS, HTTP_USER_AGENT, RAW_ROOT
from career_data.repository import CareerDataRepository


class AdapterError(ValueError):
    pass


class StructureChangedError(AdapterError):
    pass


@dataclass(frozen=True)
class RemoteDocument:
    url: str
    title: str
    applicable_year: int | None = None
    published_at: datetime | None = None


class HttpDownloader:
    def __init__(self, timeout: float = HTTP_TIMEOUT_SECONDS, retries: int = HTTP_MAX_RETRIES) -> None:
        self.timeout, self.retries = timeout, retries

    def get(self, url: str) -> bytes:
        last_error: Exception | None = None
        for attempt in range(self.retries + 1):
            try:
                response = httpx.get(url, timeout=self.timeout, follow_redirects=True,
                                     headers={"User-Agent": HTTP_USER_AGENT})
                response.raise_for_status()
                return response.content
            except (httpx.HTTPError, OSError) as exc:
                last_error = exc
                if attempt < self.retries:
                    sleep(0.5 * (attempt + 1))
        raise AdapterError(f"download failed after limited retries: {last_error}")


def normalized_header(value: Any) -> str:
    return re.sub(r"[\s\n\r：:（）()]+", "", str(value or "")).lower()


def read_tabular(path: Path) -> list[dict[str, Any]]:
    if path.stat().st_size == 0:
        raise AdapterError("empty file")
    def records_from_rows(rows: list[tuple[Any, ...] | list[Any]]) -> list[dict[str, Any]]:
        if not rows:
            return []
        sample = rows[: min(20, len(rows))]
        max_populated = max(sum(value not in (None, "") for value in row) for row in sample)
        header_index = next(
            index for index, row in enumerate(sample)
            if sum(value not in (None, "") for value in row) == max_populated
        )
        headers = [str(value or "").strip() for value in rows[header_index]]
        return [dict(zip(headers, row)) for row in rows[header_index + 1:]
                if any(value not in (None, "") for value in row)]

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        records: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            records.extend(records_from_rows(list(sheet.iter_rows(values_only=True))))
        if not records:
            raise AdapterError("empty spreadsheet")
        return records
    if path.suffix.lower() == ".xls":
        workbook = xlrd.open_workbook(str(path), on_demand=True)
        records = []
        try:
            for sheet in workbook.sheets():
                records.extend(records_from_rows([sheet.row_values(index) for index in range(sheet.nrows)]))
        finally:
            workbook.release_resources()
        if not records:
            raise AdapterError("empty spreadsheet")
        return records
    if path.suffix.lower() in {".csv", ".tsv"}:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle, delimiter=delimiter))
    raise AdapterError(f"unsupported tabular file: {path.suffix}")


def read_text_document(path: Path) -> str:
    if path.stat().st_size == 0:
        raise AdapterError("empty file")
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        with pdfplumber.open(path) as pdf:
            return "\n".join(page.extract_text() or "" for page in pdf.pages)
    if suffix in {".html", ".htm"}:
        raw = path.read_text(encoding="utf-8", errors="replace")
        raw = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", raw)
        return re.sub(r"\s+", " ", unescape(re.sub(r"(?s)<[^>]+>", "\n", raw))).replace("\ufeff", "").strip()
    return path.read_text(encoding="utf-8-sig", errors="strict").replace("\ufeff", "")


def resolve_fields(row: dict[str, Any], aliases: dict[str, tuple[str, ...]], required: set[str]) -> dict[str, Any]:
    by_header = {normalized_header(key): value for key, value in row.items()}
    result: dict[str, Any] = {}
    missing: list[str] = []
    for target, candidates in aliases.items():
        matched = next((by_header[normalized_header(name)] for name in candidates if normalized_header(name) in by_header), None)
        result[target] = matched.strip() if isinstance(matched, str) else matched
        if target in required and matched in (None, ""):
            missing.append(target)
    if missing:
        raise StructureChangedError("missing required fields/header changed: " + ", ".join(missing))
    return result


class SourceAdapter(ABC):
    code: str
    parser_version = "1.0.0"
    publisher: str
    base_url: str
    allowed_domains: tuple[str, ...]

    def __init__(self, repository: CareerDataRepository, downloader: HttpDownloader | None = None,
                 raw_root: Path | None = None) -> None:
        self.repository = repository
        self.downloader = downloader or HttpDownloader()
        self.raw_root = (raw_root or RAW_ROOT).resolve()

    def validate_source_url(self, url: str) -> None:
        host = (urlparse(url).hostname or "").lower()
        if not any(host == domain or host.endswith("." + domain) for domain in self.allowed_domains):
            raise AdapterError(f"non-official source URL rejected for {self.code}: {url}")

    def discover(self) -> list[RemoteDocument]:
        return []

    def download(self, document: RemoteDocument) -> Path:
        self.validate_source_url(document.url)
        content = self.downloader.get(document.url)
        content_hash = hashlib.sha256(content).hexdigest()
        existing = self.repository.find_document_by_hash(content_hash)
        if existing:
            existing_path = Path(existing["local_path"])
            if not existing_path.is_absolute():
                existing_path = DATA_ROOT / existing_path
            if existing_path.is_file():
                return existing_path
        suffix = Path(urlparse(document.url).path).suffix or ".bin"
        target = self.raw_root / self.code / (content_hash + suffix.lower())
        target.parent.mkdir(parents=True, exist_ok=True)
        if not target.exists():
            target.write_bytes(content)
        return target

    @abstractmethod
    def parse(self, path: Path, **context: Any) -> list[dict[str, Any]]: ...

    @abstractmethod
    def normalize(self, record: dict[str, Any], **context: Any) -> dict[str, Any]: ...

    @abstractmethod
    def validate(self, record: dict[str, Any]) -> list[str]: ...

    def persist(self, records: list[dict[str, Any]], document_id: int, source_url: str) -> tuple[int, int, int]:
        return self.repository.persist_records(self.code, records, document_id, source_url)

    def archive_local_file(self, source: Path, content_hash: str) -> Path:
        source = source.resolve()
        try:
            source.relative_to(self.raw_root)
            return source
        except ValueError:
            pass
        target = self.raw_root / self.code / f"{content_hash}{source.suffix.lower()}"
        target.parent.mkdir(parents=True, exist_ok=True)
        if not target.exists():
            shutil.copy2(source, target)
        return target

    @staticmethod
    def sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for block in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(block)
        return digest.hexdigest()

    @staticmethod
    def mime_type(path: Path) -> str:
        return mimetypes.guess_type(path.name)[0] or "application/octet-stream"
